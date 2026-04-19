"""
Fancy Serial Number Analyzer
Analyzes U.S. currency serial numbers for collectible patterns
and generates eBay-style listing titles.
"""

import argparse
import ast
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from functools import lru_cache
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False

DATE_MATCH_CATEGORY_CELEBRITY = "CELEBRITY BIRTHDAY DATASET MATCH"
DATE_MATCH_CATEGORY_HISTORICAL = "HISTORICAL DATE DATASET MATCH"
DATE_MATCH_CATEGORY_WORLD_EVENTS = "WORLD EVENT DATE DATASET MATCH"
DATE_MATCH_CATEGORY_US_HOLIDAY = "US HOLIDAY DATASET MATCH"

DEFAULT_DATASET_FILES = {
    "birthdays": "birthdays.csv",
    "historical": "World Important Dates.csv",
    "world_events": "disorder_events_sample.csv",
    "us_holidays": "us_public_holidays.csv",
    "zip_reference": "us_zip_reference.csv",
}

# Conservative OCR-like repairs when users paste noisy serial text.
OCR_DIGIT_REPAIRS = {
    "O": "0",
    "o": "0",
    "I": "1",
    "l": "1",
    "S": "5",
    "s": "5",
    "B": "8",
}


def module_base_dir():
    """
    Resolve script directory when available, otherwise current working dir.
    Useful for runtimes like Pyodide where __file__ may be unset.
    """
    module_file = globals().get("__file__")
    if module_file:
        return Path(module_file).resolve().parent
    return Path(".").resolve()


def parse_month_name(value):
    text = (value or "").strip()
    if not text:
        return None
    for fmt in ("%B", "%b"):
        try:
            return datetime.strptime(text, fmt).month
        except ValueError:
            pass
    return None


def parse_int(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def is_valid_month_day(month, day):
    try:
        datetime(2000, int(month), int(day))
        return True
    except (TypeError, ValueError):
        return False


def is_valid_date(year, month, day):
    try:
        datetime(int(year), int(month), int(day))
        return True
    except (TypeError, ValueError):
        return False


def short_text(value, limit=80):
    text = re.sub(r"\s+", " ", (value or "").strip())
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def extract_world_year(value):
    text = (value or "").strip()
    if not text:
        return None
    upper = text.upper()
    if "BC" in upper or "BCE" in upper:
        return None
    match = re.search(r"\d{1,4}", text)
    if not match:
        return None
    return int(match.group(0))


def parse_month_day_label(value):
    text = (value or "").strip().title()
    if not text:
        return None
    try:
        dt = datetime.strptime(text, "%B %d")
        return dt.month, dt.day
    except ValueError:
        return None


def parse_birthdays_entries(raw_birthdays):
    try:
        entries = ast.literal_eval(raw_birthdays)
    except (ValueError, SyntaxError):
        return []
    if isinstance(entries, list):
        return entries
    return []


def extract_birth_years(raw_birthdays):
    """
    Conservative birth-year extraction to avoid false positives.
    Only years with explicit birth indicators are accepted.
    """
    years = set()
    entries = parse_birthdays_entries(raw_birthdays)

    explicit_patterns = [
        re.compile(r"\bborn(?:\s+on)?(?:\s+[A-Za-z]+\s+\d{1,2},?)?\s+(?:in\s+)?(\d{3,4})\b", re.I),
        re.compile(r"\(b\.\s*(\d{3,4})\)", re.I),
        re.compile(r"^\s*(\d{3,4})\b"),
    ]
    range_pattern = re.compile(r"\((\d{3,4})\s*[-–]\s*(\d{2,4})\)")
    death_pattern = re.compile(r"\(d\.\s*(\d{3,4})\)", re.I)

    current_year = datetime.now().year
    for entry in entries:
        if not isinstance(entry, str):
            continue
        for pattern in explicit_patterns:
            match = pattern.search(entry)
            if not match:
                continue
            year = int(match.group(1))
            if 1000 <= year <= current_year + 1:
                years.add(year)
                break

        death_match = death_pattern.search(entry)
        range_match = range_pattern.search(entry)
        if death_match and range_match:
            start_year = int(range_match.group(1))
            end_raw = range_match.group(2)
            if len(end_raw) == 2:
                century = start_year // 100
                end_year = (century * 100) + int(end_raw)
                if end_year < start_year:
                    end_year += 100
            else:
                end_year = int(end_raw)

            death_year = int(death_match.group(1))
            age = death_year - start_year
            if end_year == death_year and 15 <= age <= 110 and 1000 <= start_year <= current_year + 1:
                years.add(start_year)

    return years


def sample_celebrity_names(raw_birthdays, max_names=3):
    entries = parse_birthdays_entries(raw_birthdays)
    if not entries:
        return []
    names = []
    for entry in entries:
        if not isinstance(entry, str):
            continue
        name = re.split(r",|\[", entry, maxsplit=1)[0].strip().strip("'\"")
        if name:
            names.append(short_text(name, 50))
        if len(names) >= max_names:
            break
    return names


class DatasetDateMatcher:
    """Indexes external date datasets for fast 8-digit serial lookups."""

    def __init__(self, max_samples_per_match=3):
        self.max_samples = max_samples_per_match
        self.exact_samples = defaultdict(lambda: defaultdict(list))
        self.exact_counts = defaultdict(lambda: defaultdict(int))
        self.birthday_by_mmdd = {}
        self.birthday_years_by_mmdd = defaultdict(set)
        self.zip_reference = {}
        self.historical_years = set()
        self.historical_embed_years = set()
        self.stats = Counter()
        self.warnings = []

    def _add_exact(self, serial_digits, category, detail):
        self.exact_counts[serial_digits][category] += 1
        bucket = self.exact_samples[serial_digits][category]
        if len(bucket) < self.max_samples:
            if detail not in bucket:
                bucket.append(detail)

    def add_date_match(self, month, day, year, category, detail):
        if not is_valid_date(year, month, day):
            return
        mmddyyyy = f"{month:02d}{day:02d}{year:04d}"
        ddmmyyyy = f"{day:02d}{month:02d}{year:04d}"
        self._add_exact(mmddyyyy, category, f"{detail} [MMDDYYYY]")
        self._add_exact(ddmmyyyy, category, f"{detail} [DDMMYYYY]")
        self.stats[f"{category}_dates_indexed"] += 1

    def load_birthdays(self, csv_path):
        if not csv_path.exists():
            self.warnings.append(f"Missing birthdays dataset: {csv_path}")
            return
        with csv_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            for row in csv.DictReader(handle):
                parsed = parse_month_day_label(row.get("Date"))
                if not parsed:
                    continue
                month, day = parsed
                if not is_valid_month_day(month, day):
                    continue
                count_val = parse_int(row.get("Count"))
                examples = sample_celebrity_names(row.get("Birthdays", ""))
                count_display = count_val if count_val is not None else "unknown"
                if examples:
                    summary = f"{datetime(2000, month, day):%B %d}: {count_display} entries (e.g., {', '.join(examples)})"
                else:
                    summary = f"{datetime(2000, month, day):%B %d}: {count_display} entries"
                self.birthday_by_mmdd[f"{month:02d}{day:02d}"] = summary
                years = extract_birth_years(row.get("Birthdays", ""))
                if years:
                    self.birthday_years_by_mmdd[f"{month:02d}{day:02d}"].update(years)
                    self.stats["celebrity_birth_years_indexed"] += len(years)
                self.stats["celebrity_days_indexed"] += 1

    def load_historical_dates(self, csv_path):
        if not csv_path.exists():
            self.warnings.append(f"Missing historical dataset: {csv_path}")
            return
        with csv_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            for row in csv.DictReader(handle):
                day = parse_int(row.get("Date"))
                month = parse_month_name(row.get("Month"))
                year = extract_world_year(row.get("Year"))
                if not (day and month and year is not None):
                    continue
                incident = short_text(row.get("Name of Incident", "Historical event"), 90)
                detail = f"{incident} ({month:02d}/{day:02d}/{year:04d})"
                self.add_date_match(month, day, year, DATE_MATCH_CATEGORY_HISTORICAL, detail)
                self.stats["historical_rows_parsed"] += 1
                if 1000 <= year <= 2100:
                    self.historical_years.add(year)
                    self.historical_embed_years.add(year)

    def load_world_events(self, csv_path):
        if not csv_path.exists():
            self.warnings.append(f"Missing world events dataset: {csv_path}")
            return
        with csv_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            for row in csv.DictReader(handle):
                raw_date = (row.get("event_date") or "").strip()
                if not raw_date:
                    continue
                try:
                    dt = datetime.strptime(raw_date, "%d %B %Y")
                except ValueError:
                    continue
                event_type = short_text(row.get("event_type", "Event"), 28)
                country = short_text(row.get("country", "Unknown country"), 24)
                location = short_text(row.get("location", "Unknown location"), 24)
                detail = f"{event_type} - {country} / {location} ({dt:%m/%d/%Y})"
                self.add_date_match(
                    dt.month, dt.day, dt.year, DATE_MATCH_CATEGORY_WORLD_EVENTS, detail
                )
                self.stats["world_event_rows_parsed"] += 1

    def load_us_holidays(self, csv_path):
        if not csv_path.exists():
            self.warnings.append(f"Missing US holidays dataset: {csv_path}")
            return
        with csv_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            for row in csv.DictReader(handle):
                raw_date = (row.get("date") or "").strip()
                name = short_text(
                    row.get("name")
                    or row.get("localName")
                    or row.get("holiday_name")
                    or "Holiday",
                    80,
                )
                if not raw_date:
                    continue
                try:
                    dt = datetime.strptime(raw_date, "%Y-%m-%d")
                except ValueError:
                    continue
                detail = f"{name} ({dt:%m/%d/%Y})"
                self.add_date_match(dt.month, dt.day, dt.year, DATE_MATCH_CATEGORY_US_HOLIDAY, detail)
                self.stats["us_holiday_rows_parsed"] += 1

    def load_zip_reference(self, csv_path):
        if csv_path.exists():
            with csv_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
                for row in csv.DictReader(handle):
                    zip_code = (row.get("zip") or row.get("postal_code") or "").strip()
                    if len(zip_code) != 5 or not zip_code.isdigit():
                        continue
                    city = short_text(row.get("city") or row.get("place_name") or "Unknown city", 30)
                    state = short_text(row.get("state") or row.get("admin_name1") or "NA", 24)
                    lat = (row.get("latitude") or "").strip()
                    lon = (row.get("longitude") or "").strip()
                    self.zip_reference[zip_code] = {
                        "city": city,
                        "state": state,
                        "latitude": lat,
                        "longitude": lon,
                    }
                    self.stats["zip_reference_rows_parsed"] += 1
            return

        # Fallback: read GeoNames-format US/US.txt (tab-delimited) when CSV is absent.
        geonames_txt = csv_path.parent / "US" / "US.txt"
        if geonames_txt.exists():
            with geonames_txt.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
                reader = csv.reader(handle, delimiter="\t")
                for row in reader:
                    if len(row) < 11:
                        continue
                    zip_code = (row[1] or "").strip()
                    if len(zip_code) != 5 or not zip_code.isdigit():
                        continue
                    if zip_code in self.zip_reference:
                        continue
                    city = short_text((row[2] or "").strip() or "Unknown city", 30)
                    state = short_text((row[3] or "").strip() or "NA", 24)
                    lat = (row[9] or "").strip()
                    lon = (row[10] or "").strip()
                    self.zip_reference[zip_code] = {
                        "city": city,
                        "state": state,
                        "latitude": lat,
                        "longitude": lon,
                    }
                    self.stats["zip_reference_rows_parsed"] += 1
            self.stats["zip_reference_source_geonames_txt"] += 1
            return

        self.warnings.append(
            f"Missing ZIP reference dataset: {csv_path} (and fallback missing: {geonames_txt})"
        )

    def load_all(self, base_dir):
        root = Path(base_dir)
        self.load_birthdays(root / DEFAULT_DATASET_FILES["birthdays"])
        self.load_historical_dates(root / DEFAULT_DATASET_FILES["historical"])
        self.load_world_events(root / DEFAULT_DATASET_FILES["world_events"])
        self.load_us_holidays(root / DEFAULT_DATASET_FILES["us_holidays"])
        self.load_zip_reference(root / DEFAULT_DATASET_FILES["zip_reference"])

    def _build_detail(self, samples, total_count):
        if not samples:
            return f"{total_count} dataset matches"
        detail = "; ".join(samples)
        if total_count > len(samples):
            detail += f"; +{total_count - len(samples)} more"
        return detail

    def _match_celebrity_full_date(self, digits):
        details = []
        seen = set()
        seen_date_keys = set()
        if len(digits) != 8 or not digits.isdigit():
            return ""

        year = int(digits[4:])
        if year < 1000 or year > datetime.now().year + 1:
            return ""

        # Strict MMDDYYYY
        mm1, dd1 = int(digits[:2]), int(digits[2:4])
        if is_valid_month_day(mm1, dd1):
            key = f"{mm1:02d}{dd1:02d}"
            years = self.birthday_years_by_mmdd.get(key, set())
            summary = self.birthday_by_mmdd.get(key)
            if summary and year in years:
                date_key = (mm1, dd1, year)
                if date_key not in seen_date_keys:
                    token = f"{summary} [MMDDYYYY year={year}]"
                    details.append(token)
                    seen.add(token)
                    seen_date_keys.add(date_key)

        # Strict DDMMYYYY
        dd2, mm2 = int(digits[:2]), int(digits[2:4])
        if is_valid_month_day(mm2, dd2):
            key = f"{mm2:02d}{dd2:02d}"
            years = self.birthday_years_by_mmdd.get(key, set())
            summary = self.birthday_by_mmdd.get(key)
            if summary and year in years:
                date_key = (mm2, dd2, year)
                if date_key not in seen_date_keys:
                    token = f"{summary} [DDMMYYYY year={year}]"
                    if token not in seen:
                        details.append(token)
                        seen.add(token)
                        seen_date_keys.add(date_key)

        return "; ".join(details[: self.max_samples])

    def match(self, digits):
        patterns = []

        exact_by_category = self.exact_counts.get(digits, {})
        for category in (
            DATE_MATCH_CATEGORY_US_HOLIDAY,
            DATE_MATCH_CATEGORY_HISTORICAL,
            DATE_MATCH_CATEGORY_WORLD_EVENTS,
        ):
            total = exact_by_category.get(category, 0)
            if not total:
                continue
            samples = self.exact_samples[digits].get(category, [])
            patterns.append((category, self._build_detail(samples, total)))

        celebrity_detail = self._match_celebrity_full_date(digits)
        if celebrity_detail:
            patterns.append((DATE_MATCH_CATEGORY_CELEBRITY, celebrity_detail))

        return patterns


DATASET_DATE_MATCHER = None
DATASET_MATCHER_CACHE = {}


def _dataset_cache_signature(base_dir):
    root = Path(base_dir).resolve()
    signatures = []
    for key in (
        "birthdays",
        "historical",
        "world_events",
        "us_holidays",
        "zip_reference",
    ):
        file_name = DEFAULT_DATASET_FILES[key]
        path = root / file_name
        if path.exists():
            stat = path.stat()
            signatures.append((file_name, stat.st_mtime_ns, stat.st_size))
        else:
            signatures.append((file_name, None, None))
    return str(root), tuple(signatures)


def configure_dataset_date_matcher(data_dir):
    global DATASET_DATE_MATCHER
    key = _dataset_cache_signature(data_dir)
    matcher = DATASET_MATCHER_CACHE.get(key)
    if matcher is None:
        matcher = DatasetDateMatcher()
        matcher.load_all(data_dir)
        DATASET_MATCHER_CACHE[key] = matcher
    DATASET_DATE_MATCHER = matcher
    return matcher


DIGIT_TO_A1Z26 = {str(i): chr(ord("A") + i - 1) for i in range(1, 10)}
ZERO_REPAIR_LETTER = "O"

WORD_LEXICON = {
    "AID", "AIDE", "BAD", "BAG", "BEAD", "BEE", "BIDE", "BOD", "BOO", "CAB",
    "CAFE", "CHIEF", "CHIDE", "COED", "COFFEE", "CODE", "DEAF", "DECADE",
    "DECO", "DECODE", "DICE", "DIE", "DIG", "DOE", "DOG", "EDGE", "EGO", "FACE",
    "FADE", "FEED", "FIB", "FIG", "FOOD", "GOAD", "GOD", "GOOD", "HAG", "HIDE",
    "HOG", "HOOD", "ICE", "IDEA", "OBOE", "OFF",
}

NAME_LEXICON = {
    "ABE", "ADA", "AIDA", "BEA", "BOB", "DEE", "DIA", "DIDI", "EDDIE",
    "GIA", "GIGI", "HEIDI",
}

BRAND_FRAGMENT_LEXICON = {
    "AEO", "AIG", "IDEO", "IHG",
}


def decode_a1z26_digitwise(digits):
    """
    Converts each single digit with A1Z26.
    1-9 => A-I, and invalid 0 is repaired to O.
    """
    repaired = []
    ignored = []
    zero_count = 0

    for d in digits:
        if d == "0":
            repaired.append(ZERO_REPAIR_LETTER)
            zero_count += 1
            continue
        letter = DIGIT_TO_A1Z26.get(d, "")
        if letter:
            repaired.append(letter)
            ignored.append(letter)

    return {
        "repaired": "".join(repaired),
        "ignore_zero": "".join(ignored),
        "zero_count": zero_count,
    }


def token_hits(decoded, lexicon):
    if not decoded:
        return []
    hits = [token for token in lexicon if token in decoded]
    hits.sort(key=lambda x: (-len(x), x))
    return hits


def readable_subsequences(decoded, min_len=3, max_len=6, limit=6):
    if not decoded:
        return []

    vowels = set("AEIOU")
    first_index = {}
    best = {}

    for start in range(len(decoded)):
        for end in range(start + min_len, min(len(decoded), start + max_len) + 1):
            sub = decoded[start:end]
            if sub in first_index:
                continue
            if len(set(sub)) == 1:
                continue

            types = [ch in vowels for ch in sub]
            if all(types) or not any(types):
                continue

            transitions = sum(types[i] != types[i + 1] for i in range(len(types) - 1))
            alternating_like = transitions >= len(sub) - 2
            repeated_pair = len(sub) >= 4 and sub[:2] == sub[2:4]

            if alternating_like or repeated_pair:
                first_index[sub] = start
                best[sub] = (len(sub), start)

    ordered = sorted(best.keys(), key=lambda s: (-best[s][0], best[s][1], s))
    return ordered[:limit]


def score_lexical_signal(words, names, brands, readable, zero_count):
    score = 0
    score += sum(4 + min(6, len(w)) for w in words[:6])
    score += sum(3 + min(5, len(n)) for n in names[:6])
    score += sum(3 + min(5, len(b)) for b in brands[:6])
    score += sum(1 + min(4, len(r) - 1) for r in readable[:6])

    if zero_count == 0:
        score += 4
    else:
        score -= min(8, zero_count * 2)

    return max(0, min(100, score))


@lru_cache(maxsize=100000)
def longest_sequential_run(digits):
    d = [int(x) for x in digits]
    best = 1
    for start in range(len(d)):
        for direction in (1, -1):
            length = 1
            while start + length < len(d) and d[start + length] == d[start] + direction * length:
                length += 1
            best = max(best, length)
    return best


def score_pattern_quality(digits, pattern_names):
    names = [p.upper() for p in pattern_names]

    def has(fragment):
        return any(fragment in n for n in names)

    counts = digit_counts(digits)
    distinct = len(counts)
    top = max(counts.values())

    repetition = min(35, (top - 1) * 6 + (8 - distinct) * 2)

    symmetry = 0
    if digits == digits[::-1]:
        symmetry += 22
    if digits[:4] == digits[4:]:
        symmetry += 16
    if digits[:2] * 4 == digits:
        symmetry += 18
    if digits[0] == digits[-1]:
        symmetry += 4
    if digits[:2] == digits[-2:]:
        symmetry += 5

    structure = 0
    seq = longest_sequential_run(digits)
    if seq >= 4:
        structure += seq * 2
    if len(set(digits[0::2])) == 1 and len(set(digits[1::2])) == 1 and digits[0] != digits[1]:
        structure += 10
    if has("TOMBSTONE"):
        structure += 8
    if has("DATASET MATCH"):
        structure += 6
    if has("BIRTHDAY / DATE NOTE"):
        structure += 6

    return max(0, min(100, repetition + symmetry + structure))


def signal_band(score):
    if score >= 85:
        return "ELITE"
    if score >= 70:
        return "HIGH"
    if score >= 55:
        return "MEDIUM-HIGH"
    if score >= 40:
        return "MEDIUM"
    return "LOW"


def join_hits(values, default="none", limit=6):
    if not values:
        return default
    return ", ".join(values[:limit])


def analyze_letter_signal(digits, pattern_names):
    decode = decode_a1z26_digitwise(digits)
    repaired = decode["repaired"]
    ignore_zero = decode["ignore_zero"]

    words = token_hits(repaired, WORD_LEXICON)
    names = token_hits(repaired, NAME_LEXICON)
    brands = token_hits(repaired, BRAND_FRAGMENT_LEXICON)
    readable = readable_subsequences(repaired)

    if ignore_zero and ignore_zero != repaired:
        for token in token_hits(ignore_zero, WORD_LEXICON):
            if token not in words:
                words.append(token)
        for token in token_hits(ignore_zero, NAME_LEXICON):
            if token not in names:
                names.append(token)
        for token in token_hits(ignore_zero, BRAND_FRAGMENT_LEXICON):
            if token not in brands:
                brands.append(token)
        for token in readable_subsequences(ignore_zero):
            if token not in readable:
                readable.append(token)

    lexical_score = score_lexical_signal(words, names, brands, readable, decode["zero_count"])
    pattern_score = score_pattern_quality(digits, pattern_names)
    high_signal_score = int(
        round(
            min(
                100,
                (pattern_score * 0.75) + (lexical_score * 0.45) + 15,
            )
        )
    )
    band = signal_band(high_signal_score)

    mapping_detail = (
        f"repaired={repaired}; ignore0={ignore_zero or '-'}; zero_repairs={decode['zero_count']}"
    )
    lexical_detail = (
        f"words=[{join_hits(words)}]; names=[{join_hits(names)}]; "
        f"brands=[{join_hits(brands)}]; readable=[{join_hits(readable)}]"
    )
    score_detail = (
        f"{high_signal_score}/100 ({band}) | lexical={lexical_score} pattern={pattern_score}"
    )
    compact_hits = (
        f"words:{join_hits(words, default='-')} | names:{join_hits(names, default='-')} | "
        f"brands:{join_hits(brands, default='-')} | readable:{join_hits(readable, default='-')}"
    )

    return {
        "mapping_detail": mapping_detail,
        "lexical_detail": lexical_detail,
        "score_detail": score_detail,
        "decoded_repaired": repaired,
        "decoded_ignore_zero": ignore_zero,
        "lexical_score": lexical_score,
        "pattern_score": pattern_score,
        "high_signal_score": high_signal_score,
        "band": band,
        "compact_hits": compact_hits,
    }

# ─────────────────────────────────────────────
#  PATTERN DETECTION
# ─────────────────────────────────────────────

def get_digits(serial):
    """Extract only numeric digits from serial string."""
    return re.sub(r'[^0-9]', '', serial)


def normalize_serial_token(token):
    """
    Normalize one serial token into collector-friendly form.
    Supports noisy input containing separators and optional '*' marker.
    """
    raw = str(token or "").strip()
    if not raw:
        return {
            "raw": raw,
            "normalized": "",
            "digits": "",
            "is_star": False,
            "status": "empty",
            "reason": "empty token",
            "corrections": [],
        }

    is_star = "*" in raw
    corrections = []
    digits = []
    invalid_letters = []
    had_separators = False

    for ch in raw:
        if ch.isdigit():
            digits.append(ch)
            continue
        if ch == "*":
            continue
        if ch in "-_/.,:; ":
            had_separators = True
            continue
        repaired = OCR_DIGIT_REPAIRS.get(ch)
        if repaired is not None:
            digits.append(repaired)
            corrections.append(f"{ch}->{repaired}")
            continue
        if ch.isalpha():
            invalid_letters.append(ch)

    digit_str = "".join(digits)
    normalized = f"{digit_str}{'*' if is_star and digit_str else ''}"

    if len(digit_str) == 8:
        status = "valid"
        reason = "ok"
    elif len(digit_str) < 8:
        status = "insufficient"
        reason = f"insufficient digits ({len(digit_str)}/8)"
    else:
        status = "too_many"
        reason = f"too many digits ({len(digit_str)}/8)"

    if invalid_letters:
        letter_note = f"ignored letters: {''.join(invalid_letters)}"
        corrections.append(letter_note)
    if had_separators:
        corrections.append("removed separators")

    return {
        "raw": raw,
        "normalized": normalized,
        "digits": digit_str,
        "is_star": is_star,
        "status": status,
        "reason": reason,
        "corrections": corrections,
    }


def parse_serial_input_text(raw_text):
    """
    Parse pasted serial input text into normalized serial values and issues.
    """
    text = str(raw_text or "")
    base_tokens = [t for t in re.split(r"[\s,;]+", text) if t.strip()]
    tokens = []
    for token in base_tokens:
        # Allow uninterrupted digit streams (no separators) by chunking into
        # 8-digit serial-sized pieces. Any trailing short chunk is surfaced
        # as an input issue by downstream normalization.
        if token.isdigit() and len(token) > 8:
            for i in range(0, len(token), 8):
                tokens.append(token[i : i + 8])
        else:
            tokens.append(token)

    parsed_tokens = []
    valid_serials = []
    valid_corrections = []
    issues = []
    seen = set()

    for token in tokens:
        parsed = normalize_serial_token(token)
        parsed_tokens.append(parsed)

        if parsed["status"] == "valid":
            serial_value = parsed["normalized"]
            key = (parsed["digits"], parsed["is_star"])
            if key not in seen:
                seen.add(key)
                valid_serials.append(serial_value)
            if parsed["raw"] != serial_value or parsed["corrections"]:
                valid_corrections.append(
                    {
                        "raw": parsed["raw"],
                        "normalized": serial_value,
                        "corrections": parsed["corrections"],
                    }
                )
        else:
            issues.append(
                {
                    "raw": parsed["raw"],
                    "normalized": parsed["normalized"],
                    "status": parsed["status"],
                    "reason": parsed["reason"],
                    "corrections": parsed["corrections"],
                }
            )

    summary = {
        "token_count": len(tokens),
        "valid_count": len(valid_serials),
        "insufficient_count": sum(1 for p in parsed_tokens if p["status"] == "insufficient"),
        "too_many_count": sum(1 for p in parsed_tokens if p["status"] == "too_many"),
        "star_count": sum(1 for p in parsed_tokens if p["status"] == "valid" and p["is_star"]),
        "corrected_count": len(valid_corrections),
    }

    return {
        "valid_serials": valid_serials,
        "tokens": parsed_tokens,
        "issues": issues,
        "corrections": valid_corrections,
        "summary": summary,
    }


def parse_serial_inputs(serial_items):
    """
    Parse a list/iterable of serial inputs. Each entry may contain many tokens.
    """
    combined = "\n".join(str(item) for item in (serial_items or []))
    return parse_serial_input_text(combined)


def count_distinct(digits):
    return len(set(digits))


def digit_counts(digits):
    return Counter(digits)


SIGNIFICANT_YEARS = {
    1492, 1776, 1789, 1812, 1861, 1865, 1914, 1918, 1929, 1939, 1941, 1945,
    1950, 1963, 1969, 1989, 2001, 2008, 2020,
}


def digit_run_segments(digits):
    segments = []
    i = 0
    while i < len(digits):
        j = i
        while j < len(digits) and digits[j] == digits[i]:
            j += 1
        segments.append((digits[i], j - i))
        i = j
    return segments


def is_valid_us_zip(zip5):
    if not (isinstance(zip5, str) and len(zip5) == 5 and zip5.isdigit()):
        return False
    value = int(zip5)
    # Conservative USPS range check.
    return 501 <= value <= 99950


def parse_mmdd(chunk4):
    if len(chunk4) != 4 or not chunk4.isdigit():
        return None
    mm, dd = int(chunk4[:2]), int(chunk4[2:])
    if is_valid_month_day(mm, dd):
        return mm, dd
    return None


def parse_ddmm(chunk4):
    if len(chunk4) != 4 or not chunk4.isdigit():
        return None
    dd, mm = int(chunk4[:2]), int(chunk4[2:])
    if is_valid_month_day(mm, dd):
        return mm, dd
    return None


# --- Solid / Near Solid ---
def check_solid(digits):
    if len(set(digits)) == 1:
        return [("SOLID SERIAL NUMBER", digits[0])]
    return []

def check_near_solid(digits):
    c = digit_counts(digits)
    for d, n in c.items():
        if n == 7:
            return [("NEAR SOLID - 7 OF A KIND", d)]
    return []


# --- Runs (N in a Row) ---
def find_runs(digits):
    results = []
    i = 0
    while i < len(digits):
        j = i
        while j < len(digits) and digits[j] == digits[i]:
            j += 1
        run_len = j - i
        if run_len >= 3:
            results.append((run_len, digits[i], i))
        i = j
    return results  # [(length, digit, start_pos), ...]

def check_runs(digits):
    patterns = []
    runs = find_runs(digits)
    for run_len, d, pos in runs:
        label = {3: "3 IN A ROW", 4: "4 IN A ROW", 5: "5 IN A ROW",
                 6: "6 IN A ROW", 7: "7 IN A ROW", 8: "8 IN A ROW"}.get(run_len)
        if label:
            patterns.append((label, d))
    return patterns


# --- N of a Kind (scattered) ---
def check_of_a_kind(digits):
    patterns = []
    c = digit_counts(digits)
    runs = find_runs(digits)
    max_run_by_digit = {}
    for run_len, d, _ in runs:
        prev = max_run_by_digit.get(d, 0)
        if run_len > prev:
            max_run_by_digit[d] = run_len
    for d, n in c.items():
        # Only report if not already represented as an equal-length solid run.
        max_run = max_run_by_digit.get(d, 0)
        if n >= 3 and max_run < n:
            label = {3: "3 OF A KIND", 4: "4 OF A KIND", 5: "5 OF A KIND",
                     6: "6 OF A KIND", 7: "7 OF A KIND", 8: "8 OF A KIND"}.get(n)
            if label:
                patterns.append((label, d))
    return patterns


# --- Binary / Trinary ---
def check_binary(digits):
    if count_distinct(digits) == 2:
        d = sorted(set(digits))
        return [("BINARY", f"{d[0]}s & {d[1]}s")]
    return []

def check_trinary(digits):
    if count_distinct(digits) == 3:
        d = sorted(set(digits))
        return [("TRINARY", f"{d[0]}s {d[1]}s & {d[2]}s")]
    return []


# --- Radar / Palindrome ---
def check_radar(digits):
    if digits == digits[::-1]:
        return [("RADAR / PALINDROME", "reads same forwards & backwards")]
    return []

def check_mini_radar(digits):
    # Inner 6 digits palindrome
    inner = digits[1:7]
    if len(inner) == 6 and inner == inner[::-1]:
        return [("MINI RADAR", "inner 6 digits palindrome")]
    return []


# --- Repeater ---
def check_repeater(digits):
    if len(digits) == 8:
        if digits[:4] == digits[4:]:
            return [("REPEATER", f"{digits[:4]} {digits[4:]}")]
    return []

def check_quad_repeater(digits):
    """2-digit group repeating 4 times: ABABABAB"""
    if len(digits) == 8:
        pair = digits[:2]
        if all(digits[i:i+2] == pair for i in range(0, 8, 2)):
            return [("QUAD REPEATER", f"{pair} x4")]
    return []

def check_double_repeater(digits):
    """Two different 2-digit pairs each appear twice: AABBAABB or ABABCDCD"""
    if len(digits) == 8:
        # Pattern: first 4 = last 4 already caught by repeater
        # Check ABCDABCD
        if digits[:4] == digits[4:]:
            return []
        # Check if digits can be split into two repeating pairs
        pairs = [digits[i:i+2] for i in range(0, 8, 2)]
        if len(set(pairs)) == 2 and pairs[0] == pairs[2] and pairs[1] == pairs[3]:
            return [("DOUBLE REPEATER", f"{pairs[0]} & {pairs[1]}")]
    return []


# --- Flipper ---
FLIP_MAP = {'0': '0', '1': '1', '6': '9', '8': '8', '9': '6'}
def check_flipper(digits):
    if all(d in FLIP_MAP for d in digits):
        flipped = ''.join(FLIP_MAP[d] for d in reversed(digits))
        if flipped.isdigit():
            return [("FLIPPER", "valid upside-down")]
    return []


# --- Ladder ---
def check_ladder(digits):
    patterns = []
    d = [int(x) for x in digits]
    # Full ascending
    if d == sorted(d) and d == list(range(d[0], d[0]+8)):
        patterns.append(("ASCENDING LADDER", "full 8-digit"))
        return patterns
    # Full descending
    if d == sorted(d, reverse=True) and d == list(range(d[0], d[0]-8, -1)):
        patterns.append(("DESCENDING LADDER", "full 8-digit"))
        return patterns
    # Partial ladders (6, 5, 4 digits)
    best = 0
    for start in range(len(d)):
        for direction in [1, -1]:
            length = 1
            while start + length < len(d) and d[start+length] == d[start] + direction*length:
                length += 1
            if length > best:
                best = length
    if best >= 6:
        patterns.append((f"{best}-DIGIT LADDER", "consecutive sequential run"))
    elif best == 5:
        patterns.append(("5-DIGIT LADDER", "consecutive sequential run"))
    elif best == 4:
        patterns.append(("4-DIGIT LADDER", "consecutive sequential run"))

    # Broken / scattered ladder (all 8 distinct digits 1-8 or 0-7 present)
    if set(digits) == set('12345678') or set(digits) == set('01234567') or set(digits) == set('23456789'):
        patterns.append(("BROKEN LADDER", "all 8 sequential digits, scattered"))
    return patterns


# --- Bookends ---
def check_bookends(digits):
    patterns = []
    # Double digit bookends
    if digits[:2] == digits[-2:]:
        patterns.append(("DOUBLE BOOKENDS", f"{digits[:2]} at each end"))
    # Triple digit bookends
    if digits[:3] == digits[-3:]:
        patterns.append(("TRIPLE DIGIT BOOKENDS", f"{digits[:3]} at each end"))
    return patterns


# --- Low Serial Number ---
def check_low_serial(digits):
    val = int(digits)
    patterns = []
    leading_zeros = len(digits) - len(digits.lstrip('0'))
    if leading_zeros >= 5:
        patterns.append(("3-DIGIT LOW SERIAL NUMBER", f"#{int(digits):,}"))
    elif leading_zeros >= 4:
        patterns.append(("4-DIGIT LOW SERIAL NUMBER", f"#{int(digits):,}"))
    elif leading_zeros >= 3:
        patterns.append(("5-DIGIT LOW SERIAL NUMBER", f"#{int(digits):,}"))
    elif val <= 999:
        patterns.append(("LOW SERIAL NUMBER", f"#{int(digits):,}"))
    return patterns


# --- Trailing Zeros ---
def check_trailing_zeros(digits):
    trailing = len(digits) - len(digits.rstrip('0'))
    if trailing >= 4:
        return [(f"TRAILING QUAD ZEROS ({trailing} zeros)", "ends in zeros")]
    elif trailing == 3:
        return [("TRAILING TRIPLE ZEROS", "ends in 000")]
    return []


# --- Leading Zeros ---
def check_leading_quad_zeros(digits):
    leading = len(digits) - len(digits.lstrip('0'))
    if leading >= 4:
        return [(f"LEADING QUAD ZEROS ({leading} zeros)", "quad 0000 prefix")]
    return []


# --- Pairs ---
def check_pairs(digits):
    # Count only adjacent, left-to-right, non-overlapping same-digit pairs.
    total_pairs = 0
    i = 0
    while i < len(digits) - 1:
        if digits[i] == digits[i + 1]:
            total_pairs += 1
            i += 2
            continue
        i += 1

    if total_pairs >= 4:
        return [("MULTIPLE PAIRS - QUAD DOUBLES", f"{total_pairs} pairs")]
    elif total_pairs == 3:
        return [("MULTIPLE PAIRS", f"3 pairs")]
    elif total_pairs == 2:
        return [("MULTIPLE PAIRS", f"2 pairs")]
    return []


def collect_zip_quick_reference(serial_results):
    """
    Build a compact, unique zip->city/state quick reference from analyzed rows.
    """
    hits = {}
    for serial_result in serial_results or []:
        serial = serial_result.get("serial", "")
        for pattern in serial_result.get("patterns", []):
            if pattern.get("pattern_name") != "ZIP CODE NOTE":
                continue
            detail = str(pattern.get("pattern_detail", ""))
            for zip_code, start, end, city, state in re.findall(
                r"(\d{5})@(\d+)-(\d+)\s+\(([^,]+),\s*([^)]+)\)", detail
            ):
                key = (zip_code, city.strip(), state.strip())
                bucket = hits.setdefault(
                    key,
                    {"zip": zip_code, "city": city.strip(), "state": state.strip(), "matches": []},
                )
                bucket["matches"].append(
                    {"serial": serial, "position": f"{start}-{end}"}
                )

    return sorted(hits.values(), key=lambda item: (item["state"], item["city"], item["zip"]))


# --- Alternator ---
def check_alternator(digits):
    if len(set(digits[0::2])) == 1 and len(set(digits[1::2])) == 1 and digits[0] != digits[1]:
        a, b = digits[0], digits[1]
        return [("ALTERNATOR", f"{a} & {b} alternating")]
    return []


# --- Birthday / Date Note ---
def check_birthday(digits):
    """Check if serial looks like a date: MMDDYYYY or YYYYMMDD."""
    patterns = []
    if len(digits) == 8:
        # MMDDYYYY
        try:
            mm, dd, yyyy = int(digits[:2]), int(digits[2:4]), int(digits[4:])
            if 1 <= mm <= 12 and 1 <= dd <= 31 and 1800 <= yyyy <= 2099:
                from calendar import month_abbr
                patterns.append(("BIRTHDAY / DATE NOTE",
                    f"{month_abbr[mm]} {dd}, {yyyy}"))
        except:
            pass
        # YYYYMMDD
        try:
            yyyy, mm, dd = int(digits[:4]), int(digits[4:6]), int(digits[6:])
            if 1 <= mm <= 12 and 1 <= dd <= 31 and 1800 <= yyyy <= 2099:
                from calendar import month_abbr
                patterns.append(("BIRTHDAY / DATE NOTE (YYYYMMDD)",
                    f"{month_abbr[mm]} {dd}, {yyyy}"))
        except:
            pass
    return patterns


# --- Double date note ---
def check_double_date_note(digits):
    if len(digits) != 8 or not digits.isdigit():
        return []

    windows = [
        ("lead", digits[:4]),
        ("mid", digits[2:6]),
        ("tail", digits[4:]),
    ]
    labels = []
    seen_dates = set()

    for where, chunk in windows:
        mmdd = parse_mmdd(chunk)
        if mmdd:
            key = (where, mmdd[0], mmdd[1])
            if key not in seen_dates:
                labels.append(f"{where}:MMDD={mmdd[0]:02d}/{mmdd[1]:02d}")
                seen_dates.add(key)
        ddmm = parse_ddmm(chunk)
        if ddmm:
            key = (where, ddmm[0], ddmm[1])
            if key not in seen_dates:
                labels.append(f"{where}:DDMM={ddmm[0]:02d}/{ddmm[1]:02d}")
                seen_dates.add(key)

    if len(labels) >= 2:
        return [("DOUBLE DATE NOTE", "; ".join(labels[:4]))]
    return []


# --- Zip code note ---
def check_zip_code_note(digits):
    if len(digits) != 8 or not digits.isdigit():
        return []
    hits = []
    for i in range(0, 4):
        chunk = digits[i:i + 5]
        if is_valid_us_zip(chunk):
            detail = f"{chunk}@{i + 1}-{i + 5}"
            if DATASET_DATE_MATCHER is not None:
                ref = DATASET_DATE_MATCHER.zip_reference.get(chunk)
                if ref:
                    detail = f"{detail} ({ref['city']}, {ref['state']})"
            hits.append(detail)
    if hits:
        return [("ZIP CODE NOTE", ", ".join(hits[:4]))]
    return []


# --- Historical date embed ---
def check_historical_date_embed(digits):
    if len(digits) != 8 or not digits.isdigit():
        return []

    years = set(SIGNIFICANT_YEARS)
    if DATASET_DATE_MATCHER is not None:
        years.update(DATASET_DATE_MATCHER.historical_embed_years)

    hits = []
    for i in range(0, 5):
        y = int(digits[i:i + 4])
        if y in years:
            hits.append(f"{y}@{i + 1}-{i + 4}")
    if hits:
        return [("HISTORICAL DATE EMBED", ", ".join(hits[:4]))]
    return []


# --- Angel number ---
def check_angel_number(digits):
    segments = digit_run_segments(digits)
    max_run = max((length for _, length in segments), default=0)
    if max_run >= 3:
        symbol = max(segments, key=lambda x: x[1])[0]
        return [("ANGEL NUMBER", f"{symbol} repeated run={max_run}")]
    return []


# --- Repeating doubles ---
def check_repeating_doubles(digits):
    if len(digits) != 8:
        return []
    pairs = [digits[i:i + 2] for i in range(0, 8, 2)]
    if all(pair[0] == pair[1] for pair in pairs) and len(set(pairs)) >= 2:
        return [("REPEATING DOUBLES", " ".join(pairs))]
    return []


# --- Clustered repeats ---
def check_clustered_repeats(digits):
    segments = digit_run_segments(digits)
    dense = [(d, n) for d, n in segments if n >= 2]
    if len(dense) >= 3:
        detail = ", ".join(f"{d}x{n}" for d, n in dense[:4])
        return [("CLUSTERED REPEATS", detail)]
    return []


# --- Mirrored center pattern ---
def check_mirrored_center_pattern(digits):
    if len(digits) != 8 or digits == digits[::-1]:
        return []
    hits = []
    for start in (0, 1, 2):
        chunk = digits[start:start + 6]
        if chunk == chunk[::-1]:
            hits.append(f"{chunk}@{start + 1}-{start + 6}")
    if hits:
        return [("MIRRORED CENTER PATTERN", ", ".join(hits))]
    return []


# --- Almost pattern ---
def check_almost_pattern(digits):
    if len(digits) != 8:
        return []
    d = [int(x) for x in digits]
    deltas = [d[i + 1] - d[i] for i in range(7)]
    asc_hits = sum(1 for x in deltas if x == 1)
    desc_hits = sum(1 for x in deltas if x == -1)
    best = max(asc_hits, desc_hits)
    if best >= 5 and best < 7:
        direction = "ascending" if asc_hits >= desc_hits else "descending"
        return [("ALMOST PATTERN (NEAR-MATCH SEQUENCE)", f"{direction} with single break")]
    return []


# --- CQQL ---
def check_cqql(digits):
    if len(digits) != 8:
        return []
    a, b = digits[0], digits[1]
    pair = a + b
    if digits[:6] == pair * 3 and digits[6] == a:
        if digits[7] != b and abs(int(digits[7]) - int(b)) <= 1:
            return [("CQQL", f"{digits[:6]} -> {digits[6:]} minor loop variation")]
    return []


# --- Bookend repeater ---
def check_bookend_repeater(digits):
    if len(digits) != 8:
        return []
    if digits[:2] == digits[-2:]:
        middle = digits[2:6]
        if middle[:2] == middle[2:] or (middle[0] == middle[2] and middle[1] == middle[3]):
            return [("BOOKEND REPEATER", f"{digits[:2]} ... {digits[-2:]} with internal repeat")]
    return []


# --- Radar binary ---
def check_radar_binary(digits):
    if len(digits) == 8 and digits == digits[::-1] and len(set(digits)) == 2:
        digs = sorted(set(digits))
        return [("RADAR BINARY", f"{digs[0]} & {digs[1]} palindrome")]
    return []


# --- Rotator pattern ---
def check_rotator_pattern(digits):
    if len(digits) != 8 or not all(d in FLIP_MAP for d in digits):
        return []
    rotated = "".join(FLIP_MAP[d] for d in reversed(digits))
    if rotated == digits:
        return []
    structured = (
        rotated == rotated[::-1]
        or rotated[:4] == rotated[4:]
        or len(set(rotated)) <= 3
        or longest_sequential_run(rotated) >= 4
    )
    if structured:
        return [("ROTATOR PATTERN", f"{digits} -> {rotated}")]
    return []


# --- External date dataset matches ---
def check_dataset_date_matches(digits):
    if DATASET_DATE_MATCHER is None:
        return []
    return DATASET_DATE_MATCHER.match(digits)


# --- Tombstone ---
def check_tombstone(digits):
    """Two 4-digit years embedded"""
    if len(digits) == 8:
        y1 = int(digits[:4])
        y2 = int(digits[4:])
        if 1800 <= y1 <= 2100 and 1800 <= y2 <= 2100 and y1 != y2:
            return [("TOMBSTONE / DOUBLE DATE NOTE", f"{y1} - {y2}")]
    return []


# --- Full House (poker) ---
def check_full_house(digits):
    c = digit_counts(digits)
    vals = sorted(c.values(), reverse=True)
    if len(vals) >= 2 and vals[0] == 3 and vals[1] == 2:
        triple = [d for d, n in c.items() if n == 3][0]
        pair   = [d for d, n in c.items() if n == 2][0]
        return [("FULL HOUSE", f"three {triple}s and two {pair}s")]
    if len(vals) >= 2 and vals[0] == 4 and vals[1] == 2:
        quad = [d for d, n in c.items() if n == 4][0]
        pair = [d for d, n in c.items() if n == 2][0]
        return [("FULL HOUSE - QUAD", f"four {quad}s and two {pair}s")]
    return []


# ─────────────────────────────────────────────
#  MASTER ANALYSIS FUNCTION
# ─────────────────────────────────────────────

ALL_CHECKS = [
    check_solid,
    check_near_solid,
    check_radar,
    check_mini_radar,
    check_radar_binary,
    check_repeater,
    check_quad_repeater,
    check_double_repeater,
    check_repeating_doubles,
    check_bookend_repeater,
    check_cqql,
    check_flipper,
    check_rotator_pattern,
    check_ladder,
    check_almost_pattern,
    check_mirrored_center_pattern,
    check_clustered_repeats,
    check_angel_number,
    check_alternator,
    check_birthday,
    check_double_date_note,
    check_zip_code_note,
    check_historical_date_embed,
    check_dataset_date_matches,
    check_tombstone,
    check_full_house,
    check_binary,
    check_trinary,
    check_runs,
    check_of_a_kind,
    check_bookends,
    check_low_serial,
    check_trailing_zeros,
    check_leading_quad_zeros,
    check_pairs,
]

def analyze_serial(serial_input):
    """
    Returns list of dicts:
    {serial, digits, pattern_name, pattern_detail}
    """
    raw = serial_input.strip()
    is_star_note = "*" in raw
    digits = get_digits(raw)
    if len(digits) != 8:
        return [{"serial": raw, "digits": digits,
                 "pattern_name": "INVALID (need 8 digits)", "pattern_detail": ""}]

    found = []
    seen = set()
    for check in ALL_CHECKS:
        results = check(digits)
        for name, detail in results:
            if name not in seen:
                seen.add(name)
                found.append({"serial": raw, "digits": digits,
                               "pattern_name": name, "pattern_detail": detail})

    if is_star_note and "STAR NOTE / ERROR NOTE" not in seen:
        found.append(
            {
                "serial": raw,
                "digits": digits,
                "pattern_name": "STAR NOTE / ERROR NOTE",
                "pattern_detail": "Contains '*' marker",
            }
        )
        seen.add("STAR NOTE / ERROR NOTE")

    if not found:
        found.append({"serial": raw, "digits": digits,
                      "pattern_name": "NO NOTABLE PATTERN", "pattern_detail": ""})
    return found


PATTERN_VALUE_RULES = [
    ("SOLID SERIAL NUMBER", 100),
    ("STAR NOTE / ERROR NOTE", 95),
    ("RADAR / PALINDROME", 92),
    ("RADAR BINARY", 90),
    ("TOMBSTONE / DOUBLE DATE NOTE", 88),
    ("CELEBRITY BIRTHDAY DATASET MATCH", 87),
    ("HISTORICAL DATE DATASET MATCH", 86),
    ("WORLD EVENT DATE DATASET MATCH", 85),
    ("US HOLIDAY DATASET MATCH", 85),
    ("BIRTHDAY / DATE NOTE", 84),
    ("REPEATER", 82),
    ("QUAD REPEATER", 82),
    ("DOUBLE REPEATER", 81),
    ("FULL HOUSE - QUAD", 80),
    ("FULL HOUSE", 78),
    ("NEAR SOLID - 7 OF A KIND", 77),
    ("BINARY", 76),
    ("TRINARY", 72),
    ("ALTERNATOR", 70),
    ("BOOKEND REPEATER", 68),
    ("BOOKENDS", 66),
    ("DOUBLE BOOKENDS", 67),
    ("TRIPLE DIGIT BOOKENDS", 68),
    ("ZIP CODE NOTE", 65),
    ("HISTORICAL DATE EMBED", 65),
    ("ANGEL NUMBER", 64),
    ("CLUSTERED REPEATS", 62),
    ("REPEATING DOUBLES", 61),
    ("ROTATOR PATTERN", 60),
    ("FLIPPER", 58),
    ("ASCENDING LADDER", 57),
    ("DESCENDING LADDER", 57),
    ("DIGIT LADDER", 56),
    ("BROKEN LADDER", 55),
    ("HIGH-SIGNAL SCORE", 54),
    ("LEXICAL SIGNAL MATCHES", 52),
    ("A1Z26 LETTER MAPPING", 50),
    ("3-DIGIT LOW SERIAL NUMBER", 49),
    ("4-DIGIT LOW SERIAL NUMBER", 48),
    ("5-DIGIT LOW SERIAL NUMBER", 47),
    ("LOW SERIAL NUMBER", 46),
    ("TRAILING QUAD ZEROS", 44),
    ("LEADING QUAD ZEROS", 44),
    ("NO NOTABLE PATTERN", 5),
]


def pattern_value_score(pattern_name, pattern_detail=""):
    name = (pattern_name or "").upper()
    detail = (pattern_detail or "").upper()

    for fragment, score in PATTERN_VALUE_RULES:
        if fragment in name:
            if fragment == "HIGH-SIGNAL SCORE":
                match = re.search(r"(\d{1,3})/100", detail)
                if match:
                    parsed = max(0, min(100, int(match.group(1))))
                    return max(score, int(round(0.35 * parsed + 25)))
            return score

    run_match = re.search(r"(\d)\s+IN A ROW", name)
    if run_match:
        run_len = int(run_match.group(1))
        return min(75, 45 + run_len * 4)

    of_kind_match = re.search(r"(\d)\s+OF A KIND", name)
    if of_kind_match:
        of_kind = int(of_kind_match.group(1))
        return min(74, 42 + of_kind * 4)

    if "INVALID" in name:
        return 0
    return 40


def sort_patterns_by_value(patterns):
    with_scores = []
    for pattern in patterns:
        copied = dict(pattern)
        copied["pattern_value_score"] = pattern_value_score(
            copied.get("pattern_name", ""),
            copied.get("pattern_detail", ""),
        )
        copied["pattern_confidence"] = pattern_confidence(copied.get("pattern_name", ""))
        with_scores.append(copied)

    with_scores.sort(
        key=lambda p: (p["pattern_value_score"], p.get("pattern_name", "")),
        reverse=True,
    )
    return with_scores


def pattern_confidence(pattern_name):
    name = (pattern_name or "").upper()

    if "INVALID" in name or "NO NOTABLE PATTERN" in name:
        return "LOW"

    high_fragments = (
        "STAR NOTE",
        "ERROR NOTE",
        "SOLID SERIAL NUMBER",
        "RADAR / PALINDROME",
        "RADAR BINARY",
        "REPEATER",
        "QUAD REPEATER",
        "DOUBLE REPEATER",
        "BIRTHDAY / DATE NOTE",
        "DATASET MATCH",
        "TOMBSTONE",
        "ZIP CODE NOTE",
    )
    if any(fragment in name for fragment in high_fragments):
        return "HIGH"

    return "MEDIUM"


def build_sellable_summary(serial, digits, signal, patterns):
    top_pattern = max(patterns, key=lambda p: p.get("pattern_value_score", 0), default=None)
    top_value = top_pattern.get("pattern_value_score", 0) if top_pattern else 0
    top_name = top_pattern.get("pattern_name", "") if top_pattern else ""

    dataset_bonus = 0
    for pattern in patterns:
        name = (pattern.get("pattern_name") or "").upper()
        if "DATASET MATCH" in name:
            dataset_bonus = max(dataset_bonus, 6)
        if "BIRTHDAY / DATE NOTE" in name:
            dataset_bonus = max(dataset_bonus, 5)

    star_bonus = 5 if "*" in str(serial) else 0
    diversity_bonus = min(10, len(patterns))
    sellability_score = int(
        round(
            (signal["high_signal_score"] * 0.62)
            + (top_value * 0.34)
            + diversity_bonus
            + dataset_bonus
            + star_bonus
        )
    )
    sellability_score = max(0, min(100, sellability_score))

    return {
        "serial": serial,
        "digits": digits,
        "sellability_score": sellability_score,
        "high_signal_score": signal["high_signal_score"],
        "band": signal["band"],
        "top_pattern_name": top_name,
        "top_pattern_value": top_value,
        "pattern_count": len(patterns),
        "star_note": "*" in str(serial),
    }


# ─────────────────────────────────────────────
#  EBAY TITLE GENERATOR
# ─────────────────────────────────────────────

def build_ebay_title(serial, digits, pattern_name, pattern_detail, denomination="$1", series="2021"):
    """
    Builds an eBay-style title consistent with how these notes are sold.
    Max ~80 chars is typical for eBay titles.
    """
    pn = pattern_name.upper()
    c  = digit_counts(digits)
    series = str(series or "").strip()
    if series == "2021":
        series = ""
    has_star_marker = "*" in str(serial)
    star_digits = f"{digits}*" if has_star_marker else digits

    # Helper: most common digit
    top_digit = c.most_common(1)[0][0]
    top_count = c.most_common(1)[0][1]

    # ─── Match to known eBay title formats ───
    if "STAR NOTE" in pn or "ERROR NOTE" in pn:
        return f"{denomination} STAR NOTE Fancy Serial Number {star_digits} {series} Federal Reserve Note"

    if "SOLID" in pn and "NEAR" not in pn:
        return f"SOLID SERIAL NUMBER {digits} {denomination} {series} Fancy Serial # Federal Reserve Note"

    if "NEAR SOLID" in pn:
        return f"NEAR SOLID 7 OF A KIND {top_digit}s {digits} {denomination} {series} Fancy Serial Number"

    if "RADAR" in pn and "MINI" not in pn and "BINARY" not in pn and "TRINARY" not in pn:
        return f"RADAR Note True Palindrome {denomination} Serial Number {digits} Fancy Serial Number {series}"

    if "MINI RADAR" in pn:
        return f"Mini Radar {digits} Fancy Serial Number Note Palindrome {denomination} {series}"

    if "QUAD REPEATER" in pn:
        return f"Quad Repeater {digits} Fancy Serial Number {denomination} {series} One Dollar Bill"

    if "DOUBLE REPEATER" in pn:
        return f"DOUBLE REPEATER {digits} Fancy Serial Number {denomination} {series} Federal Reserve Note"

    if "REPEATER" in pn and "DOUBLE" not in pn and "QUAD" not in pn:
        return f"REPEATER {digits} Fancy Serial Number {denomination} {series} One Dollar Bill"

    if "QUAD DOUBLES" in pn or "FOUR PAIRS" in pn:
        return f"$1 DOLLAR BILL FOUR PAIRS OF #S QUAD DOUBLES FANCY SERIAL NUMBER {digits} {series}"

    if "MULTIPLE PAIRS" in pn and "BINARY" not in pn and "TRINARY" not in pn:
        return f"Multiple Pairs Fancy Serial Number {denomination} Dollar Bill {digits} {series}"

    if "ALTERNATOR" in pn:
        return f"ALTERNATOR {digits} Fancy Serial Number {denomination} {series} Alternating Digits"

    if "FLIPPER" in pn and "BINARY" not in pn:
        return f"TRUE FLIPPER Fancy Serial Number {denomination} Dollar Bill {digits} {series}"

    if "FLIPPER" in pn and "BINARY" in pn:
        return f"BINARY TRUE FLIPPER Fancy Serial Number {denomination} Dollar Bill {digits} {series}"

    if "FULL HOUSE - QUAD" in pn:
        return f"Full House QUAD Fancy Serial Number {digits} {denomination} {series} One Dollar Bill"

    if "FULL HOUSE" in pn:
        return f"Full House Fancy Serial Number {digits} {denomination} {series} One Dollar Bill"

    if "TOMBSTONE" in pn:
        years = pattern_detail
        return f"[TOMBSTONE] {denomination} Dollar Bill {years} / Fancy Serial Number Double Date Note"

    if "DATASET MATCH" in pn:
        return f"{denomination} Bill Date-Match Fancy Serial Number {digits} {series}"

    if "HIGH-SIGNAL SCORE" in pn or "LEXICAL SIGNAL" in pn or "A1Z26 LETTER" in pn:
        return f"{denomination} Bill High-Signal Serial {digits} {series}"

    if "BIRTHDAY" in pn:
        date_str = pattern_detail
        return f"{denomination} Bill Birthday Note {date_str} Fancy Serial Number {digits}"

    if "BROKEN LADDER" in pn or "SCATTERED" in pn:
        return f"{denomination} DOLLAR BILL COMPLETE BROKEN LADDER 1-8 FANCY SERIAL NUMBER {digits} {series}"

    if "ASCENDING LADDER" in pn and "full 8" in pattern_detail:
        return f"{denomination} Bill Fancy Serial Ladder 8 Digit Ascending Ladder Run 12345678 {digits}"

    if "DESCENDING LADDER" in pn:
        return f"{denomination} Bill Fancy Serial Ladder 8 Digit Descending Ladder Run 87654321 {digits}"

    if "DIGIT LADDER" in pn or "LADDER" in pn:
        num = re.search(r'(\d+)', pn)
        n = num.group(1) if num else "6"
        return f"{denomination} Bill Fancy Serial Ladder {n} Digit Ladder Run {digits} {series}"

    if "TRAILING QUAD ZEROS" in pn or "TRAILING TRIPLE ZEROS" in pn:
        return f"Trailing Quad 0s Fancy Serial Number {digits} {denomination} {series} One Dollar Bill"

    if "3-DIGIT LOW SERIAL" in pn:
        return f"** 3 DIGIT ** {denomination} {series} FANCY SERIAL # FEDERAL RESERVE NOTE ** {digits}"

    if "4-DIGIT LOW SERIAL" in pn:
        num_val = int(digits)
        return f"FOUR DIGIT LOW SERIAL NUMBER 🔥 {num_val} Quad Zero {denomination} Bill FANCY {digits}"

    if "5-DIGIT LOW SERIAL" in pn:
        return f"[ 5 DIGIT ] {series} {denomination} FANCY SERIAL # LOW SERIAL NUMBER {digits}"

    if "LOW SERIAL" in pn:
        return f"Low Serial Number Fancy Serial Number {denomination} Dollar Bill {digits} {series}"

    if "LEADING QUAD ZEROS" in pn:
        return f"Leading Quad 0s Low Fancy Serial Number {digits} {denomination} {series} One Dollar Bill"

    # Binary combos
    if "BINARY" in pn and "NEAR SOLID" in pn:
        return f"BINARY NEAR SOLID 7 OF A KIND {top_digit}s {digits} {denomination} {series} Fancy Serial #"

    if "BINARY" in pn and "RADAR" in pn:
        return f"BINARY RADAR {digits} Fancy Serial Number {denomination} {series} Federal Reserve Note"

    if "BINARY" in pn and "REPEATER" in pn:
        digs = sorted(set(digits))
        return f"BINARY REPEATER {digits} Fancy Serial Number {denomination} {series} {digs[0]}s & {digs[1]}s"

    if "BINARY" in pn:
        digs = sorted(set(digits))
        count_str = f"{top_count} OF A KIND" if top_count >= 5 else ""
        return f"BINARY {count_str} {denomination} {series} {digits} {digs[0]}s & {digs[1]}s Fancy Serial Number".strip()

    if "TRINARY" in pn and "RADAR" in pn:
        return f"RADAR - TRINARY {digits} Fancy Serial Number {denomination} {series} Palindrome"

    if "TRINARY" in pn:
        digs = sorted(set(digits))
        count_label = f"6 OF A KIND {top_digit}s " if top_count == 6 else (f"5 OF A KIND {top_digit}s " if top_count == 5 else "")
        return f"**TRUE TRINARY** {denomination} SERIES {series} {count_label}{digits} FANCY SERIAL NUMBER"

    if "IN A ROW" in pn:
        n = re.search(r'(\d+) IN A ROW', pn)
        num = n.group(1) if n else "5"
        pos = "LEADING" if digits.startswith(top_digit * int(num)) else ("TRAILING" if digits.endswith(top_digit * int(num)) else "")
        return f"{num} In A Row {top_digit}s {pos} Fancy Serial Number {denomination} Bill {digits} {series}".strip()

    if "OF A KIND" in pn and "IN A ROW" not in pn:
        n = re.search(r'(\d+) OF A KIND', pn)
        num = n.group(1) if n else "5"
        return f"{num} Of A Kind {top_digit}s Fancy Serial Number {denomination} Dollar Bill {digits} {series}"

    if "BOOKEND" in pn:
        title = f"Bookends Fancy Serial Number {denomination} Dollar Bill {digits} {series}"
    else:
        title = f"Fancy Serial Number {denomination} Bill {digits} {series}"

    # Request: omit year 2021 from titles.
    if str(series).strip() == "2021":
        title = re.sub(r"(?<!\d)2021(?!\d)", "", title)
        title = re.sub(r"\s{2,}", " ", title).strip()
    return title


# ─────────────────────────────────────────────
#  EXCEL OUTPUT
# ─────────────────────────────────────────────

# Color palette
HDR_BG    = "1A1A2E"   # dark navy
HDR_FG    = "FFFFFF"
ROW_ALT   = "F0F4FF"   # light blue-white
ACCENT    = "16213E"
TAG_COLORS = {
    "BINARY":        ("DBEAFE", "1E3A8A"),
    "TRINARY":       ("E0F2FE", "0C4A6E"),
    "SOLID":         ("FEF9C3", "854D0E"),
    "NEAR SOLID":    ("FEF9C3", "854D0E"),
    "RADAR":         ("FCE7F3", "9D174D"),
    "REPEATER":      ("DCFCE7", "166534"),
    "LADDER":        ("FFF7ED", "9A3412"),
    "FLIPPER":       ("F3E8FF", "6B21A8"),
    "BIRTHDAY":      ("FEE2E2", "991B1B"),
    "CELEBRITY BIRTHDAY": ("FFE4E6", "9F1239"),
    "HISTORICAL DATE": ("E0F2FE", "0F172A"),
    "WORLD EVENT DATE": ("ECFCCB", "365314"),
    "US HOLIDAY DATASET MATCH": ("ECFEFF", "155E75"),
    "DOUBLE DATE NOTE": ("FEF3C7", "92400E"),
    "ZIP CODE NOTE": ("E0E7FF", "3730A3"),
    "HISTORICAL DATE EMBED": ("FFE4E6", "9F1239"),
    "ANGEL NUMBER": ("FCE7F3", "9D174D"),
    "REPEATING DOUBLES": ("DBEAFE", "1E3A8A"),
    "CLUSTERED REPEATS": ("DCFCE7", "166534"),
    "MIRRORED CENTER": ("F5F3FF", "6D28D9"),
    "ALMOST PATTERN": ("FFFBEB", "92400E"),
    "CQQL": ("ECFEFF", "155E75"),
    "BOOKEND REPEATER": ("FDF2F8", "9D174D"),
    "RADAR BINARY": ("EDE9FE", "5B21B6"),
    "ROTATOR PATTERN": ("E0F2FE", "0C4A6E"),
    "STAR NOTE": ("FEF3C7", "92400E"),
    "ERROR NOTE": ("FEF3C7", "92400E"),
    "A1Z26 LETTER": ("F0F9FF", "0C4A6E"),
    "LEXICAL SIGNAL": ("ECFEFF", "155E75"),
    "HIGH-SIGNAL SCORE": ("DCFCE7", "166534"),
    "TOMBSTONE":     ("F1F5F9", "334155"),
    "LOW SERIAL":    ("ECFDF5", "065F46"),
    "BOOKEND":       ("FFF1F2", "881337"),
    "ALTERNATOR":    ("F0FDF4", "14532D"),
    "FULL HOUSE":    ("FFFBEB", "92400E"),
    "DEFAULT":       ("FFFFFF", "111827"),
}

def tag_color(pattern_name):
    pn = pattern_name.upper()
    for key, (bg, fg) in TAG_COLORS.items():
        if key in pn:
            return bg, fg
    return TAG_COLORS["DEFAULT"]


def build_csv_fallback(rows, output_path, serial_rankings=None):
    fallback_path = Path(output_path)
    if fallback_path.suffix.lower() == ".xlsx":
        fallback_path = fallback_path.with_suffix(".csv")

    fieldnames = [
        "serial",
        "digits",
        "pattern_name",
        "pattern_confidence",
        "pattern_value_score",
        "pattern_detail",
        "ebay_title",
    ]
    with fallback_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Saved CSV fallback (openpyxl not installed): {fallback_path}")

    if serial_rankings:
        ranking_path = fallback_path.with_name(f"{fallback_path.stem}_signal_ranking.csv")
        ranking_fields = [
            "rank",
            "serial",
            "digits",
            "high_signal_score",
            "band",
            "lexical_score",
            "pattern_score",
            "a1z26_repaired",
            "a1z26_ignore0",
            "lexical_hits",
        ]
        with ranking_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=ranking_fields)
            writer.writeheader()
            writer.writerows(serial_rankings)
        print(f"Saved high-signal ranking CSV: {ranking_path}")
    return str(fallback_path)


def build_excel(rows, output_path, serial_rankings=None):
    if not OPENPYXL_AVAILABLE:
        build_csv_fallback(rows, output_path, serial_rankings=serial_rankings)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Fancy Serial Analysis"

    # ── Column definitions ──
    cols = [
        ("Serial Number",   22),
        ("Digits Only",     14),
        ("Pattern Name",    30),
        ("Confidence",      12),
        ("Pattern Value Score", 18),
        ("Pattern Detail",  36),
        ("eBay Listing Title", 72),
    ]

    thin = Side(style='thin', color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row ──
    for col_idx, (col_name, col_width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(name="Arial", bold=True, color=HDR_FG, size=10)
        cell.fill = PatternFill("solid", start_color=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 22

    # ── Data rows ──
    for row_idx, row in enumerate(rows, 2):
        alt = (row_idx % 2 == 0)
        bg, fg = tag_color(row["pattern_name"])

        values = [
            row["serial"],
            row["digits"],
            row["pattern_name"],
            row.get("pattern_confidence", ""),
            row.get("pattern_value_score", ""),
            row["pattern_detail"],
            row["ebay_title"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg if col_idx >= 3 else ("F8FAFC" if alt else "FFFFFF"))
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 6))
            cell.border = border
        ws.row_dimensions[row_idx].height = 30

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A1:G{len(rows)+1}"

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Pattern Summary")
    from collections import Counter as Ctr
    pattern_counts = Ctr(r["pattern_name"] for r in rows)

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 12

    h1 = ws2.cell(row=1, column=1, value="Pattern Name")
    h2 = ws2.cell(row=1, column=2, value="Count")
    for cell in [h1, h2]:
        cell.font = Font(name="Arial", bold=True, color=HDR_FG, size=10)
        cell.fill = PatternFill("solid", start_color=HDR_BG)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, (pattern, count) in enumerate(pattern_counts.most_common(), 2):
        bg, fg = tag_color(pattern)
        c1 = ws2.cell(row=i, column=1, value=pattern)
        c2 = ws2.cell(row=i, column=2, value=count)
        for cell in [c1, c2]:
            cell.font = Font(name="Arial", size=9, color=fg)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    if serial_rankings:
        ws3 = wb.create_sheet("Signal Ranking")
        ranking_cols = [
            ("Rank", 8),
            ("Serial", 14),
            ("Digits", 14),
            ("High Signal Score", 18),
            ("Band", 14),
            ("Lexical Score", 14),
            ("Pattern Score", 14),
            ("A1Z26 Repaired", 16),
            ("A1Z26 Ignore0", 16),
            ("Lexical Hits", 70),
        ]

        for col_idx, (col_name, col_width) in enumerate(ranking_cols, 1):
            cell = ws3.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(name="Arial", bold=True, color=HDR_FG, size=10)
            cell.fill = PatternFill("solid", start_color=HDR_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            ws3.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws3.row_dimensions[1].height = 22

        for row_idx, item in enumerate(serial_rankings, 2):
            values = [
                item["rank"],
                item["serial"],
                item["digits"],
                item["high_signal_score"],
                item["band"],
                item["lexical_score"],
                item["pattern_score"],
                item["a1z26_repaired"],
                item["a1z26_ignore0"],
                item["lexical_hits"],
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=9, color="111827")
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 10))
                cell.border = border
            ws3.row_dimensions[row_idx].height = 24

        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = f"A1:J{len(serial_rankings)+1}"

    wb.save(output_path)
    print(f"✅ Saved: {output_path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def run(
    serials,
    denomination="$1",
    series="2021",
    output_path="fancy_serial_analysis.xlsx",
    data_dir=None,
    include_dataset_dates=True,
):
    global DATASET_DATE_MATCHER

    if include_dataset_dates:
        dataset_root = Path(data_dir) if data_dir else module_base_dir()
        matcher = configure_dataset_date_matcher(dataset_root)
        for warning in matcher.warnings:
            print(f"WARNING: {warning}")
    else:
        DATASET_DATE_MATCHER = None

    all_rows = []
    serial_rankings = []
    sellable_serials = []
    for s in serials:
        s = s.strip()
        if not s:
            continue
        digits = get_digits(s)
        patterns = analyze_serial(s)

        if len(digits) == 8 and digits.isdigit():
            signal = analyze_letter_signal(
                digits=digits,
                pattern_names=[p["pattern_name"] for p in patterns],
            )
            patterns.extend(
                [
                    {
                        "serial": s,
                        "digits": digits,
                        "pattern_name": "A1Z26 LETTER MAPPING",
                        "pattern_detail": signal["mapping_detail"],
                    },
                    {
                        "serial": s,
                        "digits": digits,
                        "pattern_name": "LEXICAL SIGNAL MATCHES",
                        "pattern_detail": signal["lexical_detail"],
                    },
                    {
                        "serial": s,
                        "digits": digits,
                        "pattern_name": "HIGH-SIGNAL SCORE",
                        "pattern_detail": signal["score_detail"],
                    },
                ]
            )

            serial_rankings.append(
                {
                    "rank": 0,
                    "serial": s,
                    "digits": digits,
                    "high_signal_score": signal["high_signal_score"],
                    "band": signal["band"],
                    "lexical_score": signal["lexical_score"],
                    "pattern_score": signal["pattern_score"],
                    "a1z26_repaired": signal["decoded_repaired"],
                    "a1z26_ignore0": signal["decoded_ignore_zero"],
                    "lexical_hits": signal["compact_hits"],
                }
            )

        patterns = sort_patterns_by_value(patterns)

        if len(digits) == 8 and digits.isdigit():
            sellable_serials.append(
                build_sellable_summary(
                    serial=s,
                    digits=digits,
                    signal=signal,
                    patterns=patterns,
                )
            )

        for p in patterns:
            title = build_ebay_title(
                s,
                p["digits"],
                p["pattern_name"],
                p["pattern_detail"],
                denomination=denomination,
                series=series,
            )
            all_rows.append(
                {
                    "serial": s,
                    "digits": p["digits"],
                    "pattern_name": p["pattern_name"],
                    "pattern_confidence": p.get("pattern_confidence", ""),
                    "pattern_detail": p["pattern_detail"],
                    "pattern_value_score": p["pattern_value_score"],
                    "ebay_title": title,
                }
            )

    serial_rankings.sort(
        key=lambda r: (r["high_signal_score"], r["lexical_score"], r["pattern_score"]),
        reverse=True,
    )
    for idx, item in enumerate(serial_rankings, 1):
        item["rank"] = idx

    sellable_serials.sort(
        key=lambda x: (
            x["sellability_score"],
            x["high_signal_score"],
            x["top_pattern_value"],
            x["pattern_count"],
        ),
        reverse=True,
    )

    if all_rows:
        build_excel(all_rows, output_path, serial_rankings=serial_rankings)
    else:
        print("No valid serial rows to write to Excel.")
    return all_rows


def analyze_serials_for_web(
    serials,
    denomination="$1",
    series="2021",
    data_dir=None,
    include_dataset_dates=False,
    raw_input_text=None,
):
    """
    Analyze serials and return JSON-serializable results for web UIs.
    No files are created.
    """
    global DATASET_DATE_MATCHER

    if include_dataset_dates:
        dataset_root = Path(data_dir) if data_dir else module_base_dir()
        matcher = configure_dataset_date_matcher(dataset_root)
        dataset_warnings = list(matcher.warnings)
    else:
        DATASET_DATE_MATCHER = None
        dataset_warnings = []

    if raw_input_text is not None:
        parse_report = parse_serial_input_text(raw_input_text)
    else:
        parse_report = parse_serial_inputs(serials)

    cleaned_serials = parse_report["valid_serials"]

    all_rows = []
    serial_rankings = []
    serial_results = []
    top_sellable_serials = []

    for serial in cleaned_serials:
        digits = get_digits(serial)
        patterns = analyze_serial(serial)

        signal = analyze_letter_signal(
            digits=digits,
            pattern_names=[p["pattern_name"] for p in patterns],
        )
        patterns.extend(
            [
                {
                    "serial": serial,
                    "digits": digits,
                    "pattern_name": "A1Z26 LETTER MAPPING",
                    "pattern_detail": signal["mapping_detail"],
                },
                {
                    "serial": serial,
                    "digits": digits,
                    "pattern_name": "LEXICAL SIGNAL MATCHES",
                    "pattern_detail": signal["lexical_detail"],
                },
                {
                    "serial": serial,
                    "digits": digits,
                    "pattern_name": "HIGH-SIGNAL SCORE",
                    "pattern_detail": signal["score_detail"],
                },
            ]
        )

        patterns = sort_patterns_by_value(patterns)

        serial_rankings.append(
            {
                "rank": 0,
                "serial": serial,
                "digits": digits,
                "high_signal_score": signal["high_signal_score"],
                "band": signal["band"],
                "lexical_score": signal["lexical_score"],
                "pattern_score": signal["pattern_score"],
                "a1z26_repaired": signal["decoded_repaired"],
                "a1z26_ignore0": signal["decoded_ignore_zero"],
                "lexical_hits": signal["compact_hits"],
            }
        )

        top_sellable_serials.append(
            build_sellable_summary(
                serial=serial,
                digits=digits,
                signal=signal,
                patterns=patterns,
            )
        )

        serial_rows = []
        for pattern in patterns:
            row = {
                "serial": serial,
                "digits": pattern["digits"],
                "pattern_name": pattern["pattern_name"],
                "pattern_confidence": pattern.get("pattern_confidence", ""),
                "pattern_value_score": pattern["pattern_value_score"],
                "pattern_detail": pattern["pattern_detail"],
                "ebay_title": build_ebay_title(
                    serial,
                    pattern["digits"],
                    pattern["pattern_name"],
                    pattern["pattern_detail"],
                    denomination=denomination,
                    series=series,
                ),
            }
            all_rows.append(row)
            serial_rows.append(row)

        serial_results.append(
            {
                "serial": serial,
                "pattern_count": len(serial_rows),
                "patterns": serial_rows,
            }
        )

    serial_rankings.sort(
        key=lambda r: (r["high_signal_score"], r["lexical_score"], r["pattern_score"]),
        reverse=True,
    )
    for idx, item in enumerate(serial_rankings, 1):
        item["rank"] = idx

    top_sellable_serials.sort(
        key=lambda x: (
            x["sellability_score"],
            x["high_signal_score"],
            x["top_pattern_value"],
            x["pattern_count"],
        ),
        reverse=True,
    )

    dataset_stats = dict(DATASET_DATE_MATCHER.stats) if DATASET_DATE_MATCHER is not None else {}
    zip_quick_reference = collect_zip_quick_reference(serial_results)

    return {
        "serial_count": len(cleaned_serials),
        "pattern_row_count": len(all_rows),
        "serial_rankings": serial_rankings,
        "serial_results": serial_results,
        "dataset_warnings": dataset_warnings,
        "dataset_stats": dataset_stats,
        "input_summary": parse_report["summary"],
        "input_issues": parse_report["issues"],
        "input_corrections": parse_report.get("corrections", []),
        "normalized_serials": cleaned_serials,
        "top_sellable_serials": top_sellable_serials[:20],
        "zip_quick_reference": zip_quick_reference,
    }


def load_serials_from_file(file_path, return_report=False):
    text = Path(file_path).read_text(encoding="utf-8", errors="replace")
    report = parse_serial_input_text(text)
    if return_report:
        return report
    return report["valid_serials"]


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Analyze 8-digit serial numbers for fancy patterns plus external date datasets."
        )
    )
    parser.add_argument(
        "--serials-file",
        help="Path to a text/csv file containing serial numbers.",
    )
    parser.add_argument(
        "--serial",
        action="append",
        default=[],
        help="Single serial number (repeat this flag to add multiple).",
    )
    parser.add_argument(
        "--output",
        default="fancy_serial_analysis.xlsx",
        help="Output Excel file path.",
    )
    parser.add_argument(
        "--denomination",
        default="$1",
        help="Denomination label used in generated titles.",
    )
    parser.add_argument(
        "--series",
        default="2021",
        help="Series label used in generated titles.",
    )
    parser.add_argument(
        "--data-dir",
        default=str(module_base_dir()),
        help="Directory containing birthdays.csv, World Important Dates.csv, and disorder_events_sample.csv.",
    )
    parser.add_argument(
        "--no-dataset-dates",
        action="store_true",
        help="Disable external dataset date matching.",
    )
    parser.add_argument(
        "--json-output",
        help="Optional path to write JSON analysis output (web schema).",
    )
    parser.add_argument(
        "--json-only",
        action="store_true",
        help="Only write JSON output (skip Excel/CSV generation).",
    )
    parser.add_argument(
        "--json-indent",
        type=int,
        default=2,
        help="Indentation used for JSON output (default: 2).",
    )
    return parser.parse_args()


def collect_serial_inputs(args):
    reports = []
    if args.serials_file:
        reports.append(load_serials_from_file(args.serials_file, return_report=True))
    if args.serial:
        reports.append(parse_serial_inputs(args.serial))

    if not reports:
        return parse_serial_input_text("")

    merged_valid = []
    merged_corrections = []
    merged_issues = []
    seen = set()

    for report in reports:
        merged_issues.extend(report["issues"])
        merged_corrections.extend(report.get("corrections", []))
        for serial in report["valid_serials"]:
            digits = get_digits(serial)
            is_star = "*" in serial
            key = (digits, is_star)
            if key in seen:
                continue
            seen.add(key)
            merged_valid.append(serial)

    return {
        "valid_serials": merged_valid,
        "issues": merged_issues,
        "corrections": merged_corrections,
        "summary": {
            "token_count": sum(r["summary"]["token_count"] for r in reports),
            "valid_count": len(merged_valid),
            "insufficient_count": sum(r["summary"]["insufficient_count"] for r in reports),
            "too_many_count": sum(r["summary"]["too_many_count"] for r in reports),
            "star_count": sum(r["summary"]["star_count"] for r in reports),
            "corrected_count": len(merged_corrections),
        },
    }


def main():
    args = parse_args()
    parse_report = collect_serial_inputs(args)
    serials = parse_report["valid_serials"]
    if not serials:
        raise SystemExit(
            "No 8-digit serial numbers found. Provide --serials-file and/or --serial."
        )

    summary = parse_report["summary"]
    print(
        "Input normalization: "
        f"tokens={summary['token_count']}, "
        f"valid={summary['valid_count']}, "
        f"insufficient={summary['insufficient_count']}, "
        f"too_many={summary['too_many_count']}, "
        f"star_notes={summary['star_count']}, "
        f"corrected={summary['corrected_count']}"
    )
    if parse_report["issues"]:
        print("Input issues:")
        for issue in parse_report["issues"][:20]:
            print(
                f" - '{issue['raw']}' -> '{issue['normalized'] or '-'}': {issue['reason']}"
            )
        if len(parse_report["issues"]) > 20:
            print(f" - ... {len(parse_report['issues']) - 20} more issue(s)")
    if parse_report.get("corrections"):
        print("Input corrections:")
        for item in parse_report["corrections"][:20]:
            note = ", ".join(item["corrections"]) if item["corrections"] else "normalized"
            print(f" - '{item['raw']}' -> '{item['normalized']}' ({note})")
        if len(parse_report["corrections"]) > 20:
            print(f" - ... {len(parse_report['corrections']) - 20} more correction(s)")

    rows = []
    if not args.json_only:
        rows = run(
            serials=serials,
            denomination=args.denomination,
            series=args.series,
            output_path=args.output,
            data_dir=args.data_dir,
            include_dataset_dates=not args.no_dataset_dates,
        )

        output_display = args.output
        if not OPENPYXL_AVAILABLE and Path(args.output).suffix.lower() == ".xlsx":
            output_display = str(Path(args.output).with_suffix(".csv"))

        print(f"Analyzed {len(serials)} serials -> {len(rows)} pattern rows")
        print(f"Output file: {output_display}")

    if args.json_output or args.json_only:
        web_payload = analyze_serials_for_web(
            serials=serials,
            denomination=args.denomination,
            series=args.series,
            data_dir=args.data_dir,
            include_dataset_dates=not args.no_dataset_dates,
        )
        rendered_json = json.dumps(
            web_payload,
            indent=max(0, args.json_indent),
            ensure_ascii=False,
        )
        if args.json_output:
            json_path = Path(args.json_output)
            json_path.write_text(rendered_json, encoding="utf-8")
            print(f"JSON output: {json_path}")
        elif args.json_only:
            print(rendered_json)

    if DATASET_DATE_MATCHER is not None:
        celeb_days = DATASET_DATE_MATCHER.stats.get("celebrity_days_indexed", 0)
        hist_rows = DATASET_DATE_MATCHER.stats.get("historical_rows_parsed", 0)
        event_rows = DATASET_DATE_MATCHER.stats.get("world_event_rows_parsed", 0)
        holiday_rows = DATASET_DATE_MATCHER.stats.get("us_holiday_rows_parsed", 0)
        zip_rows = DATASET_DATE_MATCHER.stats.get("zip_reference_rows_parsed", 0)
        print(
            "Date datasets indexed: "
            f"celebrity day rows={celeb_days}, "
            f"historical rows={hist_rows}, "
            f"world event rows={event_rows}, "
            f"us holiday rows={holiday_rows}, "
            f"zip reference rows={zip_rows}"
        )


if __name__ == "__main__":
    main()
